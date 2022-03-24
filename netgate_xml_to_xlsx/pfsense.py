"""PfSense class."""
# Copyright © 2022 Appropriate Solutions, Inc. All rights reserved.

import argparse
from collections import OrderedDict
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Border, Font, NamedStyle, PatternFill, Side
import xmltodict

from .plugins.base_plugin import BasePlugin
from .plugins.support.elements import (
    sanitize_xml,
)
from .plugin_tools import discover_plugins

from .spreadsheet import write_ss_row, sheet_header, sheet_footer


class PfSense:
    """Handle all pfSense parsing and conversion."""

    def __init__(self, args: argparse.Namespace, in_filename: str) -> None:
        """
        Initialize and load XML.

        Technically a bit too much work to do in an init (since it can fail).
        """
        self.args = args
        self.in_file = (in_path := Path(in_filename))
        self.raw_xml: dict = {}
        self.pfsense: dict = {}
        self.workbook: Workbook = Workbook()

        # ss_filename is expected to be overwritten by
        self.ss_output_path = self._get_output_path(in_path)
        self._init_styles()
        self.default_alignment = Alignment(wrap_text=True, vertical="top")

        self.plugins = discover_plugins()

        self._load()

    def _get_output_path(self, in_path: Path) -> Path:
        """Generate output path based on args and in_filename."""
        out_path = self.args.output_dir / Path(f"{in_path.name}.xlsx")
        return out_path

    def _init_styles(self) -> None:
        """Iniitalized worksheet styles."""
        xlsx_header_font = Font(name="Calibri", size=16, italic=True, bold=True)
        xlsx_body_font = Font(name="Calibri", size=16)
        xlsx_footer_font = Font(name="Calibri", size=12, italic=True)

        body_border = Border(
            bottom=Side(border_style="dotted", color="00000000"),
            top=Side(border_style="dotted", color="00000000"),
            left=Side(border_style="dotted", color="00000000"),
            right=Side(border_style="dotted", color="00000000"),
        )

        alignment = Alignment(wrap_text=True, vertical="top")

        header = NamedStyle(name="header")
        header.alignment = alignment
        header.fill = PatternFill(
            "lightTrellis", fgColor="00339966"
        )  # fgColor="000000FF")  #fgColor="0000FF00")
        header.font = xlsx_header_font
        header.border = Border(
            bottom=Side(border_style="thin", color="00000000"),
            top=Side(border_style="thin", color="00000000"),
            left=Side(border_style="dotted", color="00000000"),
            right=Side(border_style="dotted", color="00000000"),
        )

        normal = NamedStyle(name="normal")
        normal.alignment = alignment
        normal.border = body_border
        normal.fill = PatternFill("solid", fgColor="FFFFFFFF")
        normal.font = xlsx_body_font

        footer = NamedStyle("footer")
        footer.alignment = Alignment(wrap_text=False, vertical="top")
        footer.border = body_border
        normal.fill = PatternFill("solid", fgColor="FFFFFFFF")
        footer.font = xlsx_footer_font

        self.workbook.add_named_style(header)
        self.workbook.add_named_style(normal)
        self.workbook.add_named_style(footer)

    def _load(self) -> OrderedDict:
        """Load and parse Netgate xml firewall configuration.

        Return pfsense keys.
        """
        self.raw_xml = self.in_file.read_text(encoding="utf-8")
        data = xmltodict.parse(self.raw_xml)
        self.pfsense = data["pfsense"]

    def _write_sheet(
        self,
        *,
        sheet_name: str,
        field_names: list[str],
        column_widths: list[int],
        rows: list[list],
    ):
        sheet = self.workbook.create_sheet(sheet_name)
        sheet_header(sheet, field_names, column_widths)

        # Define starting row num in case there are no rows to display.
        row_num = 2
        for row_num, row in enumerate(rows, start=row_num):
            write_ss_row(sheet, row, row_num)
        sheet_footer(sheet, row_num)

    def sanitize(self) -> None:
        """
        Sanitize the raw XML and save as original filename + '-sanitized'.

        The Netgate configuration file XML is well ordered and thus searchable via regex.
        """
        self.raw_xml = sanitize_xml(self.raw_xml)

        # Save sanitized XML
        parts = os.path.splitext(self.in_file)
        if len(parts) == 1:
            out_path = Path(f"{parts[0]}-sanitized")
        else:
            out_path = Path(f"{parts[0]}-sanitized{parts[1]}")
        out_path.write_text(self.raw_xml, encoding="utf-8")
        print(f"Sanitized file written: {out_path}.")

        # Delete the unsanitized file.
        self.in_file.unlink()
        print(f"Deleted original file: {self.in_file}.")

    def run(self, plugin_name: BasePlugin) -> None:
        """Run specific plugin and write sheet."""
        plugin = self.plugins[plugin_name]
        rows = plugin.run(self.pfsense)
        if not len(rows):
            # No data returned
            return

        self._write_sheet(
            sheet_name=plugin.display_name,
            field_names=plugin.field_names,
            column_widths=plugin.column_widths,
            rows=rows,
        )

    def save(self) -> None:
        """Delete empty first sheet and then save Workbook."""
        sheets = self.workbook.sheetnames
        del self.workbook[sheets[0]]
        out_path = self.ss_output_path
        self.workbook.save(out_path)
