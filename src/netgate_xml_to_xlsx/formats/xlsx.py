"""XLSX Format"""
# Copyright © 2022 Appropriate Solutions, Inc. All rights reserved.

import datetime

from openpyxl import Workbook
from openpyxl.styles import Border, Font, NamedStyle, PatternFill, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from netgate_xml_to_xlsx.sheetdata import SheetData

from .base_format import BaseFormat


class XlsxFormat(BaseFormat):
    def __init__(self, ctx: dict) -> None:
        self.ctx = ctx
        self.workbook = Workbook()
        self._init_styles()
        self.default_alignment = Alignment(wrap_text=True, vertical="top")
        self.sheet = None

    def start(self):
        """Initialization is sufficient."""
        pass

    def out(self, sheet_data: SheetData) -> None:
        if sheet_data is None or not sheet_data.data_rows:
            return

        self.sheet = self.workbook.create_sheet(sheet_data.sheet_name)
        self._sheet_header(self.sheet, sheet_data)

        # Define starting row num in case there are no rows to display.
        row_num = 2
        for row_num, row in enumerate(sheet_data.data_rows, start=row_num):
            self._write_row(self.sheet, row, row_num)

        self._sheet_footer(self.sheet, row_num)

    def finish(self) -> None:
        """Delete empty first sheet and then save Workbook."""
        sheets = self.workbook.sheetnames
        del self.workbook[sheets[0]]
        self.workbook.save(self.ctx["output_path"])

    def _init_styles(self) -> None:
        """Iniitalized worksheet styles."""
        xlsx_header_font = Font(name="Calibri", size=16, italic=True, bold=True)
        xlsx_body_font = Font(name="Calibri", size=16)
        xlsx_footer_font = Font(name="Calibri", size=12, italic=True)

        body_border = Border(
            bottom=Side(border_style="dotted", color="00000000"),
            top=Side(border_style="dotted", color="00000000"),
            left=Side(border_style="dotted", color="00000000"),
            right=Side(border_style="dotted", color="00000000"),
        )

        alignment = Alignment(wrap_text=True, vertical="top")

        header = NamedStyle(name="header")
        header.alignment = alignment
        header.fill = PatternFill(
            "lightTrellis", fgColor="00339966"
        )  # fgColor="000000FF")  #fgColor="0000FF00")
        header.font = xlsx_header_font
        header.border = Border(
            bottom=Side(border_style="thin", color="00000000"),
            top=Side(border_style="thin", color="00000000"),
            left=Side(border_style="dotted", color="00000000"),
            right=Side(border_style="dotted", color="00000000"),
        )

        normal = NamedStyle(name="normal")
        normal.alignment = alignment
        normal.border = body_border
        normal.fill = PatternFill("solid", fgColor="FFFFFFFF")
        normal.font = xlsx_body_font

        footer = NamedStyle("footer")
        footer.alignment = Alignment(wrap_text=False, vertical="top")
        footer.border = body_border
        normal.fill = PatternFill("solid", fgColor="FFFFFFFF")
        footer.font = xlsx_footer_font

        self.workbook.add_named_style(header)
        self.workbook.add_named_style(normal)
        self.workbook.add_named_style(footer)

    def _sheet_header(self, sheet: Worksheet, sheet_data: SheetData) -> None:
        """Write header row then set the column widths."""
        self._write_row(sheet, sheet_data.header_row, 1, "header")

        for column_number, width in enumerate(sheet_data.column_widths, start=1):
            column_letter = get_column_letter(column_number)
            sheet.column_dimensions[column_letter].width = width

    def _write_row(
        self, sheet: Worksheet, row: list, row_num: int, style_name: str = "normal"
    ) -> None:
        """
        Write a row into the spreadsheet.

        Args:
            row: A list of values to write into the row.

            row_increment: Number of rows to increment in spreadsheet before writing.

            style_name: Named XLSX style.

        Always increment the row before writing.

        """
        for column_number, value in enumerate(row, start=1):
            column_letter = get_column_letter(column_number)
            coordinate = f"{column_letter}{row_num}"
            sheet[coordinate] = value
            sheet[coordinate].style = style_name

    def _sheet_footer(self, sheet: Worksheet, row_number: int) -> None:
        """Write footer information on each sheet."""
        now = datetime.datetime.now().strftime("%Y-%m-%dT%H:%M")
        run_date = f"Run date: {now}"

        self._write_row(sheet, [run_date], row_number + 1, style_name="footer")
